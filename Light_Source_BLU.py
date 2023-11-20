from PySide6.QtWidgets import QWidget, QVBoxLayout, QLabel, QHBoxLayout, QGridLayout, \
    QFormLayout, QLineEdit, QTabWidget, QTableWidgetItem, QTableWidget, QSizePolicy, QFrame, \
    QPushButton, QAbstractItemView,QComboBox,QPushButton,QCheckBox,QDialog,QFileDialog
from PySide6.QtGui import QKeyEvent,QColor,QPalette
from PySide6.QtCore import Qt
from PySide6.QtCharts import QChart
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import sqlite3
from Setting import *
import chardet
import openpyxl
import pandas as pd

class Light_Source_BLU_button(QWidget):
    def __init__(self):
        super().__init__()

        # 實例化
        self.table = QTableWidget()
        # 設定默認值
        for row, i in enumerate(range(380,801)):
            item = QTableWidgetItem(str(i))
            item.setBackground(QColor(173, 216, 230))  # 設置背景顏色為淺藍色
            item.setTextAlignment(Qt.AlignCenter)  # 設置文本居中對齊
            self.table.setItem(row, 0, item)
        self.import_data_button = QPushButton("Import_data")
        self.export_data_button = QPushButton("Export_data")

        self.Light_Source_BLU_button_layout = QGridLayout()
        self.Light_Source_BLU_button_layout.addWidget(self.import_data_button,0,0)
        self.Light_Source_BLU_button_layout.addWidget(self.export_data_button,0,1)
        self.Light_Source_BLU_button_layout .addWidget(self.table,1,0,1,2)

        self.setLayout(self.Light_Source_BLU_button_layout)


        # 連接功能
        # 連接匯入按鈕的槽函數
        self.import_data_button.clicked.connect(self.loadExcelData)
        self.export_data_button.clicked.connect(self.exportExcelData)

    def loadExcelData(self):
        # path = "F:\Program-learning\pycharmlearing\Side_project\OPT-color-pyside\測試用頻譜.xlsx"
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        file_name, _ = QFileDialog.getOpenFileName(
            self, "選擇檔案", "", "Excel Files (*.xlsx *.xls);;Text Files (*.txt);;All Files (*)",
            options=options
        )

        if not file_name:
            return
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active


        self.table.setRowCount(sheet.max_row)
        self.table.setColumnCount(sheet.max_column)
        list_values = list(sheet.values)
        self.table.setHorizontalHeaderLabels(list_values[0])
        row_index = 0
        for value_tuple in list_values[1:]:
            print(value_tuple)
            col_index = 0
            for value in value_tuple:
                self.table.setItem(row_index,col_index,QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1
        # 創建或連接到 SQLite 資料庫
        db_path = "blu_database.db"
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # 檢查表是否存在，如果不存在就創建它
        cursor.execute("CREATE TABLE IF NOT EXISTS blu_data ({});".format(
            ', '.join(f'col_{i} TEXT' for i in range(sheet.max_column))
        ))

        # 匯入資料
        for row_values in sheet.iter_rows(min_row=2, values_only=True):
            cursor.execute("INSERT INTO blu_data VALUES ({});".format(
                ', '.join('?' for _ in row_values)
            ), row_values)

        # 提交變更
        conn.commit()

        # 關閉連線
        conn.close()
        # 更新 SQLite 資料庫路徑
        self.db_path = db_path

    def exportExcelData(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", "", "Excel Files (*.xlsx);;All Files (*)", options=options
        )

        if file_name:
            wb = openpyxl.Workbook()
            ws = wb.active

            # 匯出標題
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                if header_item is not None:
                    ws.cell(row=1, column=col + 1, value=header_item.text())

            # 匯出表格資料
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item is not None:
                        ws.cell(row=row + 2, column=col + 1, value=item.text())

            # 儲存 Excel 檔案
            wb.save(f"{file_name}.xlsx")

# class Light_Source_BLU_Table(QTableWidget):
#     def __init__(self):
#         super().__init__()
#
#         self.setColumnCount(16)
#         self.setHorizontalHeaderLabels(["波長", "項目"])
#         # 添加初始的行
#         self.setRowCount(500)
#
#
#         # 設定默認值
#         for row, i in enumerate(range(380,801)):
#             item = QTableWidgetItem(str(i))
#             item.setBackground(QColor(173, 216, 230))  # 設置背景顏色為淺藍色
#             item.setTextAlignment(Qt.AlignCenter)  # 設置文本居中對齊
#             self.setItem(row, 0, item)
#
#         # path = "F:\Program-learning\pycharmlearing\Side_project\OPT-color-pyside\測試用頻譜.xlsx"
#         # workbook = openpyxl.load_workbook(path)
#         # sheet = workbook.active
#         #
#         # self.setRowCount(sheet.max_row)
#         # self.setColumnCount(sheet.max_column)
#         # list_values = list(sheet.values)
#         # self.setHorizontalHeaderLabels(list_values[0])
#         # row_index = 0
#         # for value_tuple in list_values[1:]:
#         #     print(value_tuple)
#         #     col_index = 0
#         #     for value in value_tuple:
#         #         self.setItem(row_index, col_index, QTableWidgetItem(str(value)))
#         #         col_index += 1
#         #     row_index += 1
#
#
#         # 設置表格可編輯
#         self.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked)
#
#         # Set Background
#         # self.setStyleSheet("background-color: lightblue;")
#
#
#
#             # 創建 SQLite 數據庫，保存項目名稱和頻譜值
#             #self.create_database(lines)
#
#     def create_database(self, data):
#         conn = sqlite3.connect('your_database.db')
#         cursor = conn.cursor()
#
#         # 創建表
#         cursor.execute('''CREATE TABLE IF NOT EXISTS spectrum_data
#                           (wavelength INTEGER, item TEXT)''')
#
#         # 插入數據
#         for line in data:
#             wavelength, item = line.strip().split('\t')
#             cursor.execute("INSERT INTO spectrum_data VALUES (?, ?)", (int(wavelength), item))
#
#         conn.commit()
#         conn.close()
#
#     def loadExcelData(self):
#         path = "F:\Program-learning\pycharmlearing\Side_project\OPT-color-pyside\測試用頻譜.xlsx"
#         workbook = openpyxl.load_workbook(path)
#         sheet = workbook.active


