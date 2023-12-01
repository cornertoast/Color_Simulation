from PySide6.QtWidgets import QWidget, QVBoxLayout, QLabel, QHBoxLayout, QGridLayout, \
    QFormLayout, QLineEdit, QTabWidget, QTableWidgetItem, QTableWidget, QSizePolicy, QFrame, \
    QPushButton, QAbstractItemView,QComboBox,QPushButton,QCheckBox,QDialog,QFileDialog,QMessageBox,\
    QInputDialog,QHeaderView

from PySide6.QtGui import QKeyEvent,QColor,QPalette,QStandardItem,QKeySequence,QShortcut
from PySide6.QtWidgets import QApplication
from PySide6.QtCore import Qt
from PySide6.QtCharts import QChart
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import sqlite3
from Setting import *
import chardet
import openpyxl
import pandas as pd

class EditableHeader(QHeaderView):
    def __init__(self, orientation, parent):
        super().__init__(orientation, parent)
        self.setSectionsMovable(True)
        self.setSectionsClickable(True)

        # 添加 QLineEdit 來實現編輯功能
        self.edit_line = QLineEdit(self)
        self.edit_line.hide()
        self.edit_line.returnPressed.connect(self.commitEdit)

    def mouseDoubleClickEvent(self, event):
        index = self.logicalIndexAt(event.pos())
        self.edit_line.setGeometry(self.sectionViewportPosition(index), 0, self.sectionSize(index), self.height())
        self.edit_line.setText(self.model().headerData(index, self.orientation()))
        self.edit_line.show()
        self.edit_line.setFocus()

    def commitEdit(self):
        index = self.logicalIndexAt(self.edit_line.pos())
        self.model().setHeaderData(index, self.orientation(), self.edit_line.text())
        self.edit_line.hide()

class CIE_Spectrum(QWidget):
    def __init__(self):
        super().__init__()

        # 實例化
        self.table = QTableWidget()
        self.table.setColumnCount(16)
        # self.table.setHorizontalHeaderLabels(["波長", "項目"])
        # 添加初始的行
        self.table.setRowCount(500)
        # 使用自定義的表頭
        self.table.setHorizontalHeader(EditableHeader(self.table.horizontalHeader().orientation(), self.table))

        # 建立 QShortcut 以處理 Delete 鍵
        self.delete_shortcut = QShortcut("Delete", self)
        self.delete_shortcut.activated.connect(self.deleteSelectedCells)

        # # 設定默認值
        # for row, i in enumerate(range(380,800)):
        #     item = QTableWidgetItem(str(i))
        #     item.setBackground(QColor(173, 216, 230))  # 設置背景顏色為淺藍色
        #     item.setTextAlignment(Qt.AlignCenter)  # 設置文本居中對齊
        #     self.table.setItem(row, 0, item)
        self.import_data_button = QPushButton("Import_excel")
        self.export_data_button = QPushButton("Export_excel")
        self.add_column_button = QPushButton("Add_column")
        self.create_data_button = QPushButton("Create_Table")
        self.table_delete_button = QPushButton("Delete_table")
        # SelectQcombobox
        self.select_db_table = QComboBox()
        self.updateTableComboBox()  # 初始化時更新 ComboBox 選項
        self.select_db_table.currentIndexChanged.connect(self.tableSelectionChanged)


        self.CIE_Spectrum_layout = QGridLayout()
        self.CIE_Spectrum_layout.addWidget(self.import_data_button, 0, 0)
        self.CIE_Spectrum_layout.addWidget(self.export_data_button, 0, 1)
        self.CIE_Spectrum_layout.addWidget(self.add_column_button, 0, 2)
        self.CIE_Spectrum_layout.addWidget(self.create_data_button, 1, 0)
        self.CIE_Spectrum_layout.addWidget(self.select_db_table, 1, 1)
        self.CIE_Spectrum_layout.addWidget(self.table_delete_button, 1, 2)
        self.CIE_Spectrum_layout .addWidget(self.table, 2, 0, 1, 3)

        self.setLayout(self.CIE_Spectrum_layout)

        self.LoadDataBase()


        # 連接功能
        # 連接匯入按鈕的槽函數
        self.import_data_button.clicked.connect(self.loadExcelData)
        self.export_data_button.clicked.connect(self.exportExcelData)
        self.add_column_button.clicked.connect(self.addColumn)
        self.create_data_button.clicked.connect(self.createDatabaseFromTable)
        self.table_delete_button.clicked.connect(self.delete_table)

    def loadExcelData(self):
        # path = "F:\Program-learning\pycharmlearing\Side_project\OPT-color-pyside\測試用頻譜.xlsx"
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        file_name, _ = QFileDialog.getOpenFileName(
            self, "選擇檔案", "", "Excel Files (*.xlsx *.xls);;Text Files (*.txt);;All Files (*)",
            options=options
        )

        if not file_name:
            return
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active


        self.table.setRowCount(sheet.max_row)
        self.table.setColumnCount(sheet.max_column)
        list_values = list(sheet.values)
        self.table.setHorizontalHeaderLabels(list_values[0])
        row_index = 0
        # # 使用 list_values[0] 開始取得標題
        # header_row_values = list_values[0]
        # header_row_str = ', '.join(f'"{header}" TEXT' for header in header_row_values)

        for value_tuple in list_values[1:]:
            #print(value_tuple)
            if len(value_tuple) <= self.table.columnCount():  # 檢查欄數是否超過預期
                col_index = 0
                for value in value_tuple:
                    self.table.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                    col_index += 1
            row_index += 1
        # # 創建或連接到 SQLite 資料庫
        # db_path = "blu_database.db"
        # conn = sqlite3.connect(db_path)
        # cursor = conn.cursor()
        #
        # # 在 create table 語句中使用 header_row_str
        # cursor.execute(f'CREATE TABLE IF NOT EXISTS blu_data ({header_row_str});')
        #
        # # 匯入資料
        # for row_values in sheet.iter_rows(min_row=2, values_only=True):
        #     cursor.execute("INSERT INTO blu_data VALUES ({});".format(
        #         ', '.join('?' for _ in row_values)
        #     ), row_values)
        #
        # # 提交變更
        # conn.commit()
        #
        # # 關閉連線
        # conn.close()
        # 更新 SQLite 資料庫路徑
        # self.db_path = db_path

    def createDatabaseFromTable(self):
        # 使用 QInputDialog 取得使用者輸入的 table name
        table_name, ok = QInputDialog.getText(self, "輸入 Table 名稱", "請輸入 Table 名稱:")

        if not ok or not table_name:
            # 使用者取消或未輸入名稱，結束函數
            return

        # 創建或連接到 SQLite 資料庫
        conn = sqlite3.connect("CIE_spectrum.db")
        cursor = conn.cursor()

        # 取得表格的標題
        header_items = [self.table.horizontalHeaderItem(i).text() if self.table.horizontalHeaderItem(
            i) is not None else f'Column{i}' for i in range(self.table.columnCount())]

        # 檢查標題是否唯一
        if len(set(header_items)) != len(header_items):
            # 如果有重複，可以自行處理，例如在重複的標題後面加上編號
            header_items = [f'{header}_{i}' for i, header in enumerate(header_items)]

        header_str = ', '.join(f'"{header}" TEXT' for header in header_items)

        # 創建資料表
        cursor.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({header_str});')

        # 匯入資料
        for row in range(self.table.rowCount()):
            row_values = [self.table.item(row, col).text() if self.table.item(row, col) is not None else '' for col in
                          range(self.table.columnCount())]
            placeholder = ', '.join('?' for _ in row_values)
            cursor.execute(f'INSERT INTO "{table_name}" VALUES ({placeholder});', row_values)

        # 提交變更
        conn.commit()

        # 關閉連線
        conn.close()
        self.updateTableComboBox()
    def delete_table(self):
        # 取得現有的資料表
        conn = sqlite3.connect("CIE_spectrum.db")
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        conn.close()

        # 將元組轉換為字串列表
        table_names = [table[0] for table in tables]

        # 讓使用者選擇要刪除的資料表
        table_name, ok = QInputDialog.getItem(self, "選擇要刪除的 Table", "選擇要刪除的 Table", table_names, 0, False)

        if ok and table_name:
            # 確認使用者的選擇
            confirm_message = f"確定要刪除資料表 {table_name} 嗎？"
            reply = QMessageBox.question(self, "確認", confirm_message, QMessageBox.Yes | QMessageBox.No,
                                         QMessageBox.No)

            if reply == QMessageBox.Yes:
                # 使用者確認後，執行刪除
                conn = sqlite3.connect("CIE_spectrum.db")
                cursor = conn.cursor()
                cursor.execute(f"DROP TABLE IF EXISTS '{table_name}';")
                conn.commit()
                conn.close()

                # 刷新資料表下拉選單
                self.updateTableComboBox()

                QMessageBox.information(self, "成功", f"成功刪除資料表 {table_name}", QMessageBox.Ok)

    def exportExcelData(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", "", "Excel Files (*.xlsx);;All Files (*)", options=options
        )

        if file_name:
            wb = openpyxl.Workbook()
            ws = wb.active

            # 匯出標題
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                if header_item is not None:
                    ws.cell(row=1, column=col + 1, value=header_item.text())

            # 匯出表格資料
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item is not None:
                        ws.cell(row=row + 2, column=col + 1, value=item.text())

            # 儲存 Excel 檔案
            wb.save(f"{file_name}.xlsx")

    def addColumn(self):
        current_column_count = self.table.columnCount()

        # 新增表格欄位
        self.table.insertColumn(current_column_count)

        # # 創建或連接到 SQLite 資料庫
        # db_path = "blu_1e.db"
        # conn = sqlite3.connect(db_path)
        # cursor = conn.cursor()
        #
        # # 取得目前資料庫表格的欄位數
        # cursor.execute("PRAGMA table_info(table_data);")
        # current_db_column_count = len(cursor.fetchall())
        #
        # # 確保資料庫表格欄位與表格相同
        # if current_column_count > current_db_column_count:
        #     for i in range(current_db_column_count, current_column_count):
        #         # 修改資料庫表格，假設欄位名稱為 Column{i}
        #         cursor.execute(f'ALTER TABLE table_data ADD COLUMN Column{i} TEXT;')
        #
        # # 提交變更
        # conn.commit()
        #
        # # 關閉連線
        # conn.close()
        #
        # # 更新 SQLite 資料庫路徑
        # self.db_path = db_path

    # def keyPressEvent(self, event):
    #     super().keyPressEvent(event)
    #     if event.key() == Qt.Key_C and (event.modifiers() & Qt.ControlModifier):
    #         self.copied_cells = sorted(self.table.selectedIndexes())
    #     elif event.key() == Qt.Key_V and (event.modifiers() & Qt.ControlModifier):
    #         r = self.table.currentRow() - self.copied_cells[0].row()
    #         c = self.table.currentColumn() - self.copied_cells[0].column()
    #         for cell in self.copied_cells:
    #             self.table.setItem(cell.row() + r, cell.column() + c, QTableWidgetItem(cell.data()))

    def keyPressEvent(self, event):
        super().keyPressEvent(event)

        if event.key() == Qt.Key.Key_C and (event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            copied_cells = sorted(self.table.selectedIndexes())

            copy_text = ''
            max_column = copied_cells[-1].column()
            for c in copied_cells:
                copy_text += self.table.item(c.row(), c.column()).text()
                if c.column() == max_column:
                    copy_text += '\n'
                else:
                    copy_text += '\t'

            QApplication.clipboard().setText(copy_text)

        if event.key() == Qt.Key_V and (event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            selection = self.table.selectedIndexes()
            if selection:
                row_anchor = selection[0].row()
                column_anchor = selection[0].column()

                clipboard = QApplication.clipboard()
                rows = clipboard.text().split('\n')
                for index_row, row in enumerate(rows):
                    values = row.split('\t')
                    for index_col ,value in enumerate(values):
                        item = QTableWidgetItem(value)
                        self.table.setItem(row_anchor + index_row , column_anchor + index_col, item)
            super().keyPressEvent()

    def deleteSelectedCells(self):
        # 獲取選擇的儲存格範圍
        selected_ranges = self.table.selectedRanges()

        # 刪除選擇的儲存格內容
        for selected_range in selected_ranges:
            for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
                for column in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                    item = self.table.item(row, column)
                    if item is not None:
                        item.setText("")  # 可以根據需要進行其他操作


    def LoadDataBase(self):
        connection = sqlite3.connect("CIE_spectrum.db")
        cursor = connection.cursor()
        # 確保表格存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='CIE1931';")
        if cursor.fetchone() is None:
            connection.close()
            return  # 如果表格不存在，直接返回

        # 獲取表格的標題
        cursor.execute("PRAGMA table_info(CIE1931);")
        header_data = cursor.fetchall()
        header_labels = [column[1] for column in header_data]

        # 設置表格的列數和標題
        self.table.setRowCount(0)  # 先清空表格
        self.table.setColumnCount(len(header_labels))
        self.table.setHorizontalHeaderLabels(header_labels)

        # 獲取表格數據
        result = connection.execute("SELECT * FROM CIE1931")

        for row_number, row_data in enumerate(result):
            self.table.insertRow(row_number)
            #print("row_number", row_number)
            for column_number, data in enumerate(row_data):
                # print("column number", column_number)
                # print("row data", row_data)
                self.table.setItem(row_number, column_number, QTableWidgetItem(str(data)))
            #print("row data", row_data)
        connection.commit()
        connection.close()

    def updateTableComboBox(self):
        # 更新 ComboBox 的選項
        # 在需要更新 ComboBox 的地方呼叫這個函數
        # 例如，當你新增了新的 table 時，呼叫 updateTableComboBox() 以更新 ComboBox
        # 連接到 SQLite 資料庫
        conn = sqlite3.connect("CIE_spectrum.db")
        cursor = conn.cursor()

        # 取得所有的 table 名稱
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()

        # 更新 ComboBox 的選項
        self.select_db_table.clear()
        for table in tables:
            self.select_db_table.addItem(table[0])

        # 關閉連線
        conn.close()

    def tableSelectionChanged(self):
        # 當 QComboBox 選擇變更時觸發的函數
        selected_table = self.select_db_table.currentText()
        if selected_table:
            # 在這裡加入相應的操作，例如從選擇的資料表中擷取資料並更新到 widget_table
            self.loadTableData(selected_table)

    def loadTableData(self, table_name):
        # 在這裡加入載入資料的程式碼，將選擇的資料表的內容更新到 widget_table
        connection = sqlite3.connect("CIE_spectrum.db")
        cursor = connection.cursor()
        # 確保表格存在
        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}';")
        if cursor.fetchone() is None:
            connection.close()
            return  # 如果表格不存在，直接返回

        # 獲取表格的標題
        cursor.execute(f"PRAGMA table_info('{table_name}');")
        header_data = cursor.fetchall()
        header_labels = [column[1] for column in header_data]

        # 設置表格的列數和標題
        self.table.setRowCount(0)  # 先清空表格
        self.table.setColumnCount(len(header_labels))
        self.table.setHorizontalHeaderLabels(header_labels)

        # 獲取表格數據
        result = connection.execute(f"SELECT * FROM '{table_name}'")  # 注意這裡使用單引號將表格名稱括起來

        for row_number, row_data in enumerate(result):
            self.table.insertRow(row_number)
            #print("row_number", row_number)
            for column_number, data in enumerate(row_data):
                # print("column number", column_number)
                # print("row data", row_data)
                self.table.setItem(row_number, column_number, QTableWidgetItem(str(data)))
            #print("row data", row_data)
        connection.commit()
        connection.close()
